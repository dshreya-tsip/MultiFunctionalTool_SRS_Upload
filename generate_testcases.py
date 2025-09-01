import os
import re
import docx
import openpyxl
import requests


# -------------------------------
# Step 1: Extract text from SRS.docx
# -------------------------------
def extract_srs_text(doc_path: str) -> str:
    doc = docx.Document(doc_path)
    return "\n".join([para.text for para in doc.paragraphs if para.text.strip()])


# -------------------------------
# Step 2: Build prompt for Claude
#   - First line MUST be: "Component: <name>"
#   - Then a blank line
#   - Then a single markdown table (no extra headers)
# -------------------------------
def build_prompt(srs_text: str) -> str:
    return (
        "Read the uploaded Software Requirements Specification (SRS.docx).\n"
        "You MUST output exactly two parts in this order:\n"
        "1) A single line in the exact format:\n"
        "   Component: <detected overall component/module/system name from the SRS>\n"
        "   (Put only this line first. No code fences, no extra text before it.)\n"
        "2) A blank line, followed immediately by a single markdown table of test cases.\n\n"

        "Generate both functional and non-functional test cases. Functional test cases should cover "
        "all described features; once completed, continue numbering with non-functional test cases "
        "(performance, usability, compatibility) WITHOUT adding any new section headers or titles. "
        "All test cases must be in ONE continuous markdown table with sequential numbering for "
        "`Test Case ID` (e.g., TC001, TC002, ...).\n\n"

        "Return the markdown table with columns exactly named:\n"
        "`Test Case ID` | `Preconditions` | `Test Condition` | `Steps with description` | "
        "`Expected Result` | `Actual Result` | `Remarks`\n\n"

        "Notes for the header block in the Excel sheet (handled by my program):\n"
        "- The line you output as 'Component: <name>' will be written into the header's Component field.\n"
        "- `MFP` will be set to 'Any' if unspecified. `Build`, `Date`, and `Target` will remain blank.\n\n"

        "SRS Content:\n" + srs_text
    )


# -------------------------------
# Step 3: Send prompt to Claude API
#   - Keep full text (component line + table)
# -------------------------------
def get_testcases_from_claude(srs_text: str) -> str:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("Missing Anthropic API key. Set ANTHROPIC_API_KEY environment variable.")

    prompt = build_prompt(srs_text)

    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }

    payload = {
        "model": "claude-3-7-sonnet-20250219",
        "max_tokens": 3000,
        "temperature": 0.3,
        "messages": [
            {"role": "user", "content": prompt}
        ],
    }

    resp = requests.post("https://api.anthropic.com/v1/messages", json=payload, headers=headers, timeout=120)
    resp.raise_for_status()
    result = resp.json()

    # Join only text blocks (Claude returns a list of content blocks)
    md_full_text = "\n".join(
        block["text"] for block in result.get("content", []) if block.get("type") == "text"
    )

    # Debug (optional)
    print("\n--- Claude Raw Output (first 1200 chars) ---\n")
    print(md_full_text[:1200])
    print("\n--------------------------------------------\n")

    return md_full_text


# -------------------------------
# Step 4: Extract "Component: <name>" from the LLM output
# -------------------------------
def extract_component(md_full_text: str) -> str:
    # Look for a line starting with 'Component:'
    m = re.search(r"(?im)^\s*Component\s*:\s*(.+?)\s*$", md_full_text)
    if m:
        return m.group(1).strip()
    return "Unknown"


# -------------------------------
# Step 5: Parse the markdown table from the LLM output
#   - Finds the first line that contains 'Test Case ID' and '|'
#   - Parses until the table ends
# -------------------------------
def parse_markdown_table(md_full_text: str):
    lines = md_full_text.splitlines()

    # Find the start of the table header (first line that looks like a header row)
    start_idx = None
    for i, line in enumerate(lines):
        if "|" in line and "Test Case ID" in line:
            start_idx = i
            break

    if start_idx is None:
        raise ValueError("Markdown table header not found in model output.")

    # Collect contiguous table lines starting at start_idx
    table_lines = []
    for line in lines[start_idx:]:
        if "|" in line:
            table_lines.append(line)
        else:
            # Stop when we hit a non-table line *after* we've started collecting
            if table_lines:
                break

    if len(table_lines) < 3:
        raise ValueError("Markdown table appears incomplete (need header, separator, and at least one row).")

    header_cells = [h.strip() for h in table_lines[0].split("|")[1:-1]]
    test_cases = []

    # Skip header + separator; then parse each row
    for row_line in table_lines[2:]:
        parts = [p.strip() for p in row_line.split("|")[1:-1]]
        if len(parts) == len(header_cells):
            test_cases.append(dict(zip(header_cells, parts)))

    if not test_cases:
        raise ValueError("No test case rows parsed from the markdown table.")

    return test_cases


# -------------------------------
# Step 6: Write into Excel template
#   - Writes header 'Component: <name>' by searching first 10 rows
#   - Sets 'MFP: Any' if an 'MFP:' label cell is found
#   - Writes table starting at row 6, columns B..H (2..8)
# -------------------------------
def fill_excel_template(test_cases, template_path: str, output_path: str, component_name: str):
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Testcases"]  # keep your original sheet name

    # Helper to set header fields robustly (merged cells supported as we overwrite the anchor cell)
    def set_header_field(label: str, value: str, search_rows: int = 10, search_cols: int = 12) -> bool:
        label_low = label.lower().rstrip(":")
        for r in range(1, search_rows + 1):
            for c in range(1, search_cols + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, str):
                    text = cell.value.strip()
                    # Match either "Label:" or "Label: existing"
                    if text.lower().startswith(label_low + ":"):
                        # Preserve label casing from the cell up to ':'
                        prefix = text.split(":", 1)[0]
                        cell.value = f"{prefix}: {value}".strip()
                        return True
        return False

    # Write Component into header (try to find an existing "Component:" cell; if not, fallback to E2)
    if not set_header_field("Component", component_name):
        ws["E2"] = f"Component: {component_name}"

    # Set MFP to 'Any' if there is an MFP label present
    if not set_header_field("MFP", "Any"):
        # Only set a fallback if the template doesn't already place it elsewhere
        # (safe no-op if E3 is not used in your template)
        if ws["E3"].value in (None, "", "MFP:", "MFP"):
            ws["E3"] = "MFP: Any"

    # Leave Build/Date/Target as they are (per your instruction)

    # Write table starting row/columns (your original mapping)
    start_row = 6  # your template places the table from row 6
    for i, tc in enumerate(test_cases):
        row = start_row + i
        ws.cell(row=row, column=2, value=tc.get("Test Case ID"))           # Col B
        ws.cell(row=row, column=3, value=tc.get("Preconditions"))          # Col C
        ws.cell(row=row, column=4, value=tc.get("Test Condition"))         # Col D
        ws.cell(row=row, column=5, value=tc.get("Steps with description")) # Col E
        ws.cell(row=row, column=6, value=tc.get("Expected Result"))        # Col F
        ws.cell(row=row, column=7, value=tc.get("Actual Result"))          # Col G
        ws.cell(row=row, column=8, value=tc.get("Remarks"))                # Col H

    wb.save(output_path)


# -------------------------------
# Main
# -------------------------------
if __name__ == "__main__":
    srs_path = "SRS.docx"
    template_path = "TestCases_Template.xlsx"
    output_path = "Generated_TestCases.xlsx"

    # 1) Read SRS
    srs_text = extract_srs_text(srs_path)

    # 2) Get full LLM output (component line + table)
    md_full = get_testcases_from_claude(srs_text)

    # 3) Extract component name
    component = extract_component(md_full)
    print(f"✅ Detected Component: {component}")

    # 4) Parse markdown table
    test_cases = parse_markdown_table(md_full)

    # 5) Fill Excel (writes header + table)
    fill_excel_template(test_cases, template_path, output_path, component)

    print(f"✅ Test cases generated successfully: {output_path}")
